#coding:utf-8
import openpyxl
import time
from openpyxl.drawing.image import Image
import os
import PIL


if not os.path.exists('./����'):
    os.mkdir('./����')

fl = os.listdir('./����/')

for fname in fl:
    deals = os.listdir('./����/{}'.format(fname))
    for deal in deals:
        if deal.endswith('.txt'):
            continue
        wb = openpyxl.load_workbook('./����/{}/{}'.format(fname, deal))
        sheet = wb.active
        row = 2
        all = 0
        while sheet['a{}'.format(str(row))].value != None:
            sum = float(sheet['d{}'.format(row)].value) * float(sheet['e{}'.format(row)].value)
            sheet['f{}'.format(row)] = '%.2f' % sum
            all += sum
            row += 1
        sheet['e{}'.format(row)] = '�ܼƣ�'
        sheet['f{}'.format(row)] = '%.2f' % all
        wb.save('./����/{}/{}'.format(fname, deal))
        wb.close()


c_width = 15
r_height = 150

t = time.strftime('%Y.%m.%d',  time.localtime())
y, m, d = t.split('.')
if not os.path.exists('./����/{}��'.format(y)):
    os.mkdir('./����/{}��'.format(y))
if not os.path.exists('./����/{}��/{}��.xlsx'.format(y, m)):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['a1'] = '����'
    sheet['b1'] = 'ͼƬ'
    sheet['c1'] = '���'
    sheet['d1'] = '����'
    sheet['e1'] = '����'
    sheet['f1'] = '�ܼ�'
    sheet['g1'] = '��ʽ'
    sheet['h1'] = '��ע'
else:
    wb = openpyxl.load_workbook('./����/{}��/{}��.xlsx'.format(y, m))
    sheet = wb.active

sheet.column_dimensions['B'].width = c_width
sheet.column_dimensions['A'].width = c_width
sheet.column_dimensions['G'].width = 30
all = 0
row = 2
while sheet['a{}'.format(row)].value != None:
    all += float(sheet['d{}'.format(row)].value) * float(sheet['e{}'.format(row)].value)
    sheet['f{}'.format(row)] = '%.2f' % (float(sheet['d{}'.format(row)].value) * float(sheet['e{}'.format(row)].value))
    row += 1

while 1:
    img = input('ͼƬ,���� q �˳�\n>')[1:-1]
    if img == '':
        break
    while not os.path.exists(img):
        img = input('ͼƬ�����ڣ������´���\n>')[1:-1]
    kuanhao = input('���\n>')
    price = float(input('����\n>'))
    price = '%.2f' % (price)
    num = float(input('����\n>'))
    num = '%.2f' % (num)
    sum = '%.2f' % (float(price) * float(num))
    model = input('��ʽ\n>')
    ad = input('��ע\n>')

    img = Image(img)
    sheet.row_dimensions[row].height = r_height
    img.width, img.height = (100, 200)
    sheet.add_image(img, 'b{}'.format(row))
    sheet['a{}'.format(row)] = t
    sheet['c{}'.format(row)] = kuanhao
    sheet['d{}'.format(row)] = price
    sheet['e{}'.format(row)] = num
    sheet['f{}'.format(row)] = sum
    sheet['g{}'.format(row)] = model
    sheet['h{}'.format(row)] = ad
    all += float(sum)
    row += 1
sheet['e{}'.format(row)] = '�ܼƣ�'
sheet['f{}'.format(row)] = '%.2f' % all
print('{}���ܼ�{} Ԫ'.format(m, all))
wb.save('./����/{}��/{}��.xlsx'.format(y, m))

fl = os.listdir('./����/{}��'.format(y))
yearall = 0
for fname in fl:
    if fname.endswith('xlsx'):
        workspace = openpyxl.load_workbook('./����/{}��/{}'.format(y, fname))
        sheet = workspace.active
        j = 2
        while sheet['e{}'.format(j)].value != None:
            j += 1
        yearall += float(sheet['f{}'.format(j-1)].value)
f = open('./����/{}��/ȫ���ܼ�.txt'.format(y), 'w')
f.write('{} ��'.format(yearall))
print('ȫ���ܼ�{} Ԫ'.format(yearall))
input('���»س��˳�����\n')

f.close()
